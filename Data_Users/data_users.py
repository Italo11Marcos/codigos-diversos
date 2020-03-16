import openpyxl
from openpyxl.utils import get_column_letter,column_index_from_string

wb1 = openpyxl.load_workbook('/home/dti/Documentos/Data_Users/tabelaDadosPesquisa.xlsx')
wb2 = openpyxl.load_workbook('/home/dti/Documentos/Data_Users/teste1.xlsx')

sheet1 = wb1.get_sheet_by_name('pag1')
sheet2 = wb2.get_sheet_by_name('teste1')

names = ['aalper', 'aduscher', 'afominaya', 'AlessandradeAndrade', 'AngelicaMenchaca', 'anithoirealaigh', 'annmoney', 'ashridhar', 'asiren', 'autter', 'bdasilva1', 'BelfranCarbonellMedina', 'bnichols1', 'BrandyJoy', 'brendering', 'bvonholdt', 'CandyMorenoGarcia', 'cdean1', 'chimanleong', 'clorie1', 'cmarcelaarenasgmez', 'courtneyaccoon', 'ChristopherDanaLynn', 'crulison', 'csuharlim', 'DanielleSalcido', 'darrenrgrocke', 'djbrightsmith', 'DK15', 'dlherrerag', 'dpowers', 'dradosavljevic1', 'DrewLab', 'dyee', 'eabbey', 'ebucholz', 'Ehren', 'P9', 'P10', 'Gaby', 'gshriver', 'gruntleme', 'gspatola', 'gvanwijngaarden', 'hgarcial', 'leeadyer', 'lreyestorres', 'ihiltpold', 'IsabelMiranda', 'jboyd1', 'jcote2', 'jdavidmcinnes', 'jguillermojimnezcorts', 'jhillerislambers', 'JohnStuartReid', 'JordanChertok', 'jpauldelgado', 'jrichmond2', 'KaylaRussell', 'KCrabill', 'kateywright', 'kevinpsmith', 'kregelmann', 'KristenWaring', 'MaddieVroom', 'MadisonLevine', 'Marine_Girl', 'malmarayati', 'MattNurse', 'mfaisalhaider', 'mhowells', 'mhsmith', 'mlosee', 'MorganGoulding', 'mseymour', 'myxomop', 'mzimova', 'namodei', 'nhank', 'nliman', 'nmagdelijns', 'P18', 'nryan2', 'P20', 'pcross', 'pstap', 'RebeccaHarris1', 'RicardoMSouza', 'RodolfoGarciaContreras', 'SarahHlubik', 'SayerJi', 'scarty', 'scheijp', 'ShannonMurphy', 'SHilbourne', 'sgrundner', 'SLRichdon', 'smarcos', 'smayne', 'SongHyeInAshley', 'srigazio', 'sthompson2', 'tchapman1', 'tglass', 'TheresaSolenski', 'tmousseau', 'twicks696', 'WillJGrant', 'werdnus', 'VitaGerritsen', 'YamilLiscano', 'zaramare', 'wsimbaa']

letters = ['B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','X']
'''
for row in range(2, sheet2.max_row+ 1):
    ut = sheet2['C'+str(row)].value
    #print(ut,row)
    if ut in names:
        pos = names.index(ut)
        c = get_column_letter(pos+11)
        print(c,row)
        sheet1[str(c)+str(row)] = 1
'''
#'''
for row in range(2, sheet2.max_row+ 1):
    for i in letters:
    #ut = sheet2['B'+str(row)].value
        ut = sheet2[i+str(row)].value
        #print(i+str(row))
    #print(ut,row)
        if ut in names:
            pos = names.index(ut)
            c = get_column_letter(pos+11)
            print(c,row)
            sheet1[str(c)+str(row)] = 1
#'''   
wb1.save('CERTOTALVEZ2.xlsx')
